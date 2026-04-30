#!/usr/bin/env python3
"""
tennis_scraper.py - v4.2
"""

import re
import sys
import argparse
from datetime import datetime, timedelta
from pathlib import Path

import requests
import pandas as pd
import numpy as np
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

try:
    from colorama import init, Fore, Style
    init(autoreset=True)
except ImportError:
    class _Dummy:
        def __getattr__(self, _): return ""
    Fore = Style = _Dummy()

DEFAULT_EXCEL = "2026.xlsx"
DEFAULT_URL   = "https://www.tennisexplorer.com/miami/2026/atp-men/"
MATCHES_URL   = "https://www.tennisexplorer.com/matches/?type=atp-single"
SHEET_DATA    = "2026"
SHEET_RANKING = "Ranking"
SHEET_UPCOMING= "Upcoming"

HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/122.0.0.0"}

ATP_POINTS = {
    "Grand Slam": {
        "1st Round": 10, "2nd Round": 45, "3rd Round": 100, "4th Round": 180,
        "Quarterfinals": 360, "Semifinals": 720, "The Final": 1200, "Winner": 2000,
    },
    "Masters 1000": {
        "1st Round": 10, "2nd Round": 25, "3rd Round": 50, "4th Round": 100,
        "Quarterfinals": 200, "Semifinals": 400, "The Final": 650, "Winner": 1000,
    },
    "ATP500": {
        "1st Round": 0, "2nd Round": 20, "3rd Round": 45,
        "Quarterfinals": 100, "Semifinals": 200, "The Final": 330, "Winner": 500,
    },
    "ATP250": {
        "1st Round": 0, "2nd Round": 13, "3rd Round": 29,
        "Quarterfinals": 57, "Semifinals": 108, "The Final": 180, "Winner": 250,
    },
}

ROUND_ORDER = ["1st Round", "2nd Round", "3rd Round", "4th Round",
               "Quarterfinals", "Semifinals", "The Final"]

EXCEL_COLS = [
    "ATP", "Location", "Tournament", "Date", "Series", "Court", "Surface",
    "Round", "Best of", "Winner", "Loser", "WRank", "LRank", "WPts", "LPts",
    "W1", "L1", "W2", "L2", "W3", "L3", "W4", "L4", "W5", "L5",
    "Wsets", "Lsets", "Comment",
    "B365W", "B365L", "PSW", "PSL", "MaxW", "MaxL", "AvgW", "AvgL", "BFEW", "BFEL",
]

UPCOMING_COLS = [
    "Location", "Tournament", "Date", "Series", "Court", "Surface",
    "Round", "Player1", "Player2", "Odds1", "Odds2", "Status",
]

# ─── HELPERS ─────────────────────────────────────────────────

def hdr(msg):
    print(f"\n{Fore.CYAN}{Style.BRIGHT}{'─'*60}\n  {msg}\n{'─'*60}{Style.RESET_ALL}")
def ok(msg):   print(f"  {Fore.GREEN}✔{Style.RESET_ALL}  {msg}")
def info(msg): print(f"  {Fore.YELLOW}·{Style.RESET_ALL}  {msg}")
def warn(msg): print(f"  {Fore.YELLOW}⚠{Style.RESET_ALL}  {msg}")
def err(msg):  print(f"  {Fore.RED}✖{Style.RESET_ALL}  {msg}")

def _clean(text):
    return re.sub(r"\s+", " ", str(text or "")).strip()

def fetch_page(url):
    resp = requests.get(url, headers=HEADERS, timeout=20)
    resp.raise_for_status()
    return BeautifulSoup(resp.text, "html.parser")

def discover_current_tournament_urls(limit=None, include_challengers=False):
    soup = fetch_page(MATCHES_URL)
    discovered = []
    seen = set()
    for link in soup.find_all("a", href=True):
        href = link.get("href", "")
        text = _clean(link.get_text())
        if not re.fullmatch(r"/[^/]+/\d{4}/atp-men/", href):
            continue
        if text.lower() in {"today", "atp singles", "all matches"}:
            continue
        if not include_challengers and "chall" in text.lower():
            continue
        if any(token in text.lower() for token in ("utr", "itf", "futures", "junior")):
            continue
        full_url = f"https://www.tennisexplorer.com{href}"
        if full_url in seen:
            continue
        seen.add(full_url)
        discovered.append(full_url)
        if limit and len(discovered) >= limit:
            break
    return discovered

def strip_seed(text):
    return re.sub(r"\s*\(\d+\)\s*$", "", text).strip()

def round_name(raw):
    mapping = {
        "R128": "1st Round", "R64": "2nd Round", "R32": "3rd Round",
        "R16": "4th Round", "QF": "Quarterfinals", "SF": "Semifinals",
        "F": "The Final", "1R": "1st Round", "2R": "2nd Round", "3R": "3rd Round",
    }
    return mapping.get(raw.upper().strip(), raw)

def parse_tournament_meta(soup, url):
    meta = {"tournament": "", "location": "", "series": "ATP250",
            "surface": "Hard", "court": "Outdoor", "year": datetime.now().year}
    m = re.search(r"/(\d{4})/", url)
    if m: meta["year"] = int(m.group(1))
    m = re.search(r"\.com/([^/]+)/", url)
    if m:
        meta["location"] = m.group(1).replace("-", " ").title()
        meta["tournament"] = meta["location"]
    for m1000 in ["indian-wells","miami","monte-carlo","madrid","rome",
                  "canada","cincinnati","shanghai","paris"]:
        if m1000 in url.lower():
            meta["series"] = "Masters 1000"
            break

    # Nawierzchnia na podstawie URL turnieju
    clay_keywords = ["monte-carlo","madrid","rome","barcelona","hamburg","roland-garros",
                     "paris-open","estoril","bucharest","munich","geneva","lyon","nice"]
    grass_keywords = ["wimbledon","queens","halle","eastbourne","s-hertogenbosch","nottingham",
                      "newport","mallorca","antalya","birmingham","bad-homburg"]
    carpet_keywords = ["bercy","vienna","stockholm","basle","basel","moscow","st-petersburg"]
    url_lower = url.lower()
    if any(k in url_lower for k in clay_keywords):
        meta["surface"] = "Clay"
    elif any(k in url_lower for k in grass_keywords):
        meta["surface"] = "Grass"
    elif any(k in url_lower for k in carpet_keywords):
        meta["surface"] = "Carpet"

    return meta


# ─── DATE PARSING ───────────────────────────────────────────

def parse_date_results(cell_text, year):
    """Parsuje datę z tabeli wyników: '25.03.03:10' -> datetime"""
    m = re.match(r"(\d{1,2})\.(\d{1,2})\.", cell_text)
    if m:
        try:
            return pd.Timestamp(datetime(year, int(m.group(2)), int(m.group(1))))
        except:
            pass
    return None


def parse_date_upcoming(cell_text, year):
    """
    Parsuje datę z tabeli upcoming:
    'today, 20:00'       -> dzisiejsza data
    'tomorrow, 01:30'    -> jutrzejsza data
    '27.03. 18:00'       -> 2026-03-27
    """
    text = cell_text.lower().strip()
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    
    if text.startswith("today"):
        return pd.Timestamp(today)
    
    if text.startswith("tomorrow"):
        return pd.Timestamp(today + timedelta(days=1))
    
    # DD.MM. HH:MM lub DD.MM.
    m = re.match(r"(\d{1,2})\.(\d{1,2})\.", text)
    if m:
        try:
            return pd.Timestamp(datetime(year, int(m.group(2)), int(m.group(1))))
        except:
            pass
    
    return None


# ─── TABLE CLASSIFICATION ──────────────────────────────────

def classify_table(table):
    """
    Rozpoznaje typ tabeli:
    - 'results':  nagłówek zawiera 'S' i cyfry setów ('1','2','3')
    - 'upcoming': nagłówek zawiera 'Round' i 'H2H' (ale NIE 'S')
    - None:       inna tabela (H2H sidebar, prize money, etc.)
    """
    header_row = table.find("tr")
    if not header_row:
        return None
    
    cols = [_clean(c.get_text()).lower() for c in header_row.find_all(["th", "td"])]
    
    has_S = "s" in cols
    has_sets = any(c in cols for c in ["1", "2", "3"])
    has_round = "round" in cols or any("round" in c for c in cols)
    has_h2h = "h2h" in cols
    has_h = "h" in cols
    has_a = "a" in cols
    
    # Tabela wyników: ma S i cyfry setów
    if has_S and has_sets:
        return "results"
    
    # Tabela upcoming: ma Round + H2H + kursy (H/A)
    # ALE nie jest to sidebar H2H (ten ma TYLKO 'H2H' bez 'Round')
    if has_round and has_h2h:
        return "upcoming"
    
    # Sidebar H2H: nagłówek to tylko ['H2H', 'H2H'] - ignoruj
    if has_h2h and not has_round:
        return None
    
    return None


# ─── PARSER WYNIKÓW ─────────────────────────────────────────

def parse_results_table(table, meta):
    results = []
    year = meta["year"]
    
    player_rows = []
    for tr in table.find_all("tr"):
        links = tr.find_all("a", href=re.compile(r"/player/"))
        if not links:
            continue
        cells = tr.find_all(["td", "th"])
        texts = [_clean(c.get_text()) for c in cells]
        player_rows.append({
            "player_raw": _clean(links[0].get_text()),
            "player": strip_seed(_clean(links[0].get_text())),
            "texts": texts,
        })
    
    i = 0
    while i < len(player_rows) - 1:
        r1 = player_rows[i]
        r2 = player_rows[i + 1]
        t1 = r1["texts"]
        t2 = r2["texts"]
        p1 = r1["player"]
        p2 = r2["player"]
        
        if not p1 or not p2 or p1 == p2:
            i += 1
            continue
        
        match_date = None
        round_str = None
        
        for cell in t1[:3]:
            if re.match(r"\d{1,2}\.\d{1,2}\.", cell):
                match_date = parse_date_results(cell, year)
            if cell.upper() in ["R128","R64","R32","R16","QF","SF","F","1R","2R","3R"]:
                round_str = round_name(cell)
        
        # Kursy
        odds1 = odds2 = np.nan
        found_odds = []
        for cell in reversed(t1):
            if cell == "info": continue
            try:
                val = float(cell.replace(",", "."))
                if 1.01 < val < 100:
                    found_odds.insert(0, val)
                    if len(found_odds) >= 2: break
            except: pass
        if len(found_odds) >= 2:
            odds1, odds2 = found_odds[0], found_odds[1]
        
        # Wyniki
        def extract_scores(texts, player_raw):
            scores = []
            started = False
            for t in texts:
                if re.match(r"\d{1,2}\.\d{1,2}\.", t): continue
                if t.upper() in ["R128","R64","R32","R16","QF","SF","F","1R","2R","3R"]: continue
                if player_raw in t or strip_seed(t) == strip_seed(player_raw):
                    started = True
                    continue
                if t == "info": break
                if started:
                    if "." in t:
                        try:
                            float(t); break
                        except: pass
                    if t == "":
                        scores.append(None)
                    elif t.isdigit():
                        scores.append(int(t))
                    elif len(t) <= 3 and t[0].isdigit():
                        scores.append(int(t[0]))
                    else:
                        scores.append(None)
            return scores
        
        scores1 = extract_scores(t1, r1["player_raw"])
        scores2 = extract_scores(t2, r2["player_raw"])
        
        sets_won_1 = scores1[0] if scores1 and scores1[0] is not None else 0
        sets_won_2 = scores2[0] if scores2 and scores2[0] is not None else 0
        set_sc_1 = [s for s in scores1[1:] if s is not None] if len(scores1) > 1 else []
        set_sc_2 = [s for s in scores2[1:] if s is not None] if len(scores2) > 1 else []
        
        has_scores = len(set_sc_1) > 0 or len(set_sc_2) > 0
        anyone_won = sets_won_1 > 0 or sets_won_2 > 0
        is_completed = anyone_won and has_scores
        
        if not is_completed and has_scores and len(set_sc_1) >= 1 and len(set_sc_2) >= 1:
            is_completed = True
            sets_won_1 = sum(1 for a, b in zip(set_sc_1, set_sc_2) if a > b)
            sets_won_2 = sum(1 for a, b in zip(set_sc_1, set_sc_2) if b > a)
        
        if is_completed:
            comment = "Completed"
            all_text = " ".join(t1 + t2).lower()
            if "ret" in all_text: comment = "Retired"
            elif "w/o" in all_text or "w.o" in all_text: comment = "Walkover"
            if (sets_won_1 + sets_won_2) == 1 and comment == "Completed":
                comment = "Retired"
            
            if sets_won_1 >= sets_won_2:
                winner, loser = p1, p2
                w_sets, l_sets = sets_won_1, sets_won_2
                w_sc, l_sc = set_sc_1, set_sc_2
                avg_w, avg_l = odds1, odds2
            else:
                winner, loser = p2, p1
                w_sets, l_sets = sets_won_2, sets_won_1
                w_sc, l_sc = set_sc_2, set_sc_1
                avg_w, avg_l = odds2, odds1
            
            results.append({
                "ATP": np.nan,
                "Location": meta["location"],
                "Tournament": meta["tournament"],
                "Date": match_date if match_date else pd.NaT,
                "Series": meta["series"],
                "Court": meta["court"],
                "Surface": meta["surface"],
                "Round": round_str or "Unknown",
                "Best of": 3,
                "Winner": winner, "Loser": loser,
                "WRank": np.nan, "LRank": np.nan,
                "WPts": np.nan, "LPts": np.nan,
                "W1": w_sc[0] if len(w_sc)>0 else np.nan,
                "L1": l_sc[0] if len(l_sc)>0 else np.nan,
                "W2": w_sc[1] if len(w_sc)>1 else np.nan,
                "L2": l_sc[1] if len(l_sc)>1 else np.nan,
                "W3": w_sc[2] if len(w_sc)>2 else np.nan,
                "L3": l_sc[2] if len(l_sc)>2 else np.nan,
                "W4": w_sc[3] if len(w_sc)>3 else np.nan,
                "L4": l_sc[3] if len(l_sc)>3 else np.nan,
                "W5": w_sc[4] if len(w_sc)>4 else np.nan,
                "L5": l_sc[4] if len(l_sc)>4 else np.nan,
                "Wsets": w_sets, "Lsets": l_sets,
                "Comment": comment,
                "B365W": np.nan, "B365L": np.nan,
                "PSW": np.nan, "PSL": np.nan,
                "MaxW": np.nan, "MaxL": np.nan,
                "AvgW": avg_w, "AvgL": avg_l,
                "BFEW": np.nan, "BFEL": np.nan,
            })
        
        i += 2
    
    return results


# ─── PARSER UPCOMING ────────────────────────────────────────

def parse_upcoming_table(table, meta):
    """
    Parsuje tabelę upcoming.
    
    Struktura wiersza:
    ['today, 20:00', 'QF', 'Landaluce - Leh', 'Live streamsbwi', '0-0', '3.18', '1.35']
    
    Link: href=/match-detail/..., tekst: 'Landaluce - Lehecka'
    
    Gracze są w JEDNYM linku, rozdzieleni " - ".
    """
    upcoming = []
    year = meta["year"]
    
    rows = table.find_all("tr")[1:]  # pomiń nagłówek
    
    for row in rows:
        cells = row.find_all(["td", "th"])
        texts = [_clean(c.get_text()) for c in cells]
        
        if len(texts) < 3:
            continue
        
        # ─── GRACZE ───
        # Szukaj linku /match-detail/ z "Gracz1 (seed) - Gracz2 (seed)"
        p1 = p2 = None
        
        all_links = row.find_all("a")
        for link in all_links:
            href = link.get("href", "")
            link_text = _clean(link.get_text())
            
            # Link do meczu: "Landaluce - Lehecka"
            if "/match-detail/" in href or " - " in link_text:
                # Rozdziel po " - "
                parts = link_text.split(" - ", 1)
                if len(parts) == 2:
                    p1 = strip_seed(parts[0].strip())
                    p2 = strip_seed(parts[1].strip())
                    break
        
        # Fallback: dwa linki /player/
        if not p1 or not p2:
            player_links = row.find_all("a", href=re.compile(r"/player/"))
            if len(player_links) >= 2:
                p1 = strip_seed(_clean(player_links[0].get_text()))
                p2 = strip_seed(_clean(player_links[1].get_text()))
        
        if not p1 or not p2:
            continue
        
        # ─── DATA ───
        # Format: 'today, 20:00' / 'tomorrow, 01:30' / '27.03. 18:00'
        match_date = None
        if texts:
            match_date = parse_date_upcoming(texts[0], year)
        
        # ─── RUNDA ───
        round_str = None
        for t in texts[:3]:
            upper = t.upper().strip()
            if upper in ["R128","R64","R32","R16","QF","SF","F","1R","2R","3R"]:
                round_str = round_name(upper)
                break
        
        # ─── KURSY ───
        odds1 = odds2 = np.nan
        found_odds = []
        for t in texts:
            try:
                val = float(t.replace(",", "."))
                if 1.01 < val < 100:
                    found_odds.append(val)
            except:
                pass
        if len(found_odds) >= 2:
            # Ostatnie dwa floaty to kursy (H i A)
            odds1 = found_odds[-2]
            odds2 = found_odds[-1]
        
        upcoming.append({
            "Location": meta["location"],
            "Tournament": meta["tournament"],
            "Date": match_date if match_date else pd.NaT,
            "Series": meta["series"],
            "Court": meta["court"],
            "Surface": meta["surface"],
            "Round": round_str or "Unknown",
            "Player1": p1,
            "Player2": p2,
            "Odds1": odds1,
            "Odds2": odds2,
            "Status": "Upcoming",
        })
    
    return upcoming


# ─── SCRAPE ALL ──────────────────────────────────────────────

def scrape_all(soup, meta):
    results = []
    upcoming = []
    
    tables = soup.find_all("table")
    info(f"Znaleziono {len(tables)} tabel")
    
    for i, table in enumerate(tables):
        ttype = classify_table(table)
        
        if ttype == "results":
            info(f"Tabela {i}: WYNIKI")
            r = parse_results_table(table, meta)
            results.extend(r)
            info(f"  → {len(r)} meczów")
        
        elif ttype == "upcoming":
            info(f"Tabela {i}: UPCOMING")
            u = parse_upcoming_table(table, meta)
            upcoming.extend(u)
            info(f"  → {len(u)} meczów")
        
        else:
            # Pomijamy inne tabele (H2H sidebar, prize money, news)
            pass
    
    return results, upcoming


# ─── RANKING ─────────────────────────────────────────────────

def compute_ranking(df):
    all_names = set()
    if "Winner" in df.columns: all_names |= set(df["Winner"].dropna())
    if "Loser" in df.columns: all_names |= set(df["Loser"].dropna())
    
    rows = []
    for player in all_names:
        wins_df = df[df["Winner"] == player] if "Winner" in df.columns else pd.DataFrame()
        losses_df = df[df["Loser"] == player] if "Loser" in df.columns else pd.DataFrame()
        wins, losses = len(wins_df), len(losses_df)
        total = wins + losses
        if total == 0: continue
        
        points = 0
        tournaments = set()
        best_results = {}
        
        for _, match in wins_df.iterrows():
            series = match.get("Series", "ATP250")
            rnd = match.get("Round", "")
            tourn = match.get("Tournament", "")
            tournaments.add(tourn)
            if series in ATP_POINTS and rnd in ATP_POINTS[series]:
                points += ATP_POINTS[series][rnd]
                cur = ROUND_ORDER.index(best_results.get(tourn, "")) if best_results.get(tourn) in ROUND_ORDER else -1
                new = ROUND_ORDER.index(rnd) if rnd in ROUND_ORDER else -1
                if new > cur: best_results[tourn] = rnd
        
        for _, match in losses_df.iterrows():
            tournaments.add(match.get("Tournament", ""))
        
        best = "-"
        best_idx = -1
        for rnd in best_results.values():
            if rnd in ROUND_ORDER:
                idx = ROUND_ORDER.index(rnd)
                if idx > best_idx:
                    best_idx = idx
                    best = rnd
        
        if best == "The Final" and len(wins_df[wins_df["Round"] == "The Final"]) > 0:
            best = "Winner"
        
        rows.append({"Player": player, "Points": points, "W": wins, "L": losses,
                     "Win%": round(wins/total*100, 1), "Tournaments": len(tournaments),
                     "BestResult": best})
    
    ranking = pd.DataFrame(rows).sort_values(
        ["Points","W","Win%"], ascending=[False,False,False]
    ).reset_index(drop=True)
    ranking.index += 1
    ranking.index.name = "Rank"
    return ranking


# ─── NAME INDEX ──────────────────────────────────────────────

def build_name_index(df):
    idx = {}
    for col in ["Winner", "Loser"]:
        if col in df.columns:
            for name in df[col].dropna().unique():
                parts = str(name).split()
                if parts:
                    idx[parts[0].lower()] = name
    return idx

def normalize_name(raw, idx):
    raw = re.sub(r"\s*[\(\[].*?[\)\]]\s*", "", raw).strip()
    lower = raw.lower()
    for k, v in idx.items():
        if v.lower() == lower: return v
    first = raw.split()[0].lower() if raw.split() else ""
    if first in idx: return idx[first]
    if len(raw.split()) >= 2:
        two = " ".join(raw.split()[:2]).lower()
        for k, v in idx.items():
            if v.lower().startswith(two): return v
        # Porównanie z normalizacją myślników: "Auger Aliassime" == "Auger-Aliassime F."
        two_nohyphen = two.replace("-", " ")
        for k, v in idx.items():
            if v.lower().replace("-", " ").startswith(two_nohyphen): return v
    return raw


# ─── EXCEL WRITER ────────────────────────────────────────────

def write_excel(excel_path, new_matches, upcoming, name_idx):
    try:
        df_existing = pd.read_excel(excel_path, sheet_name=SHEET_DATA)
    except Exception as e:
        err(f"Błąd: {e}"); return None, None
    
    info(f"Istniejące: {len(df_existing)}")
    
    # Normalizuj
    for m in new_matches:
        m["Winner"] = normalize_name(str(m.get("Winner","")), name_idx)
        m["Loser"]  = normalize_name(str(m.get("Loser","")),  name_idx)
    for u in upcoming:
        u["Player1"] = normalize_name(str(u.get("Player1","")), name_idx)
        u["Player2"] = normalize_name(str(u.get("Player2","")), name_idx)
    
    # Deduplikacja wyników — pomijamy wiersze Upcoming, bo blokują dopisanie prawdziwych wyników
    df_existing_results = df_existing[df_existing.get("Comment", pd.Series(dtype=str)) != "Upcoming"] \
        if "Comment" in df_existing.columns else df_existing

    def match_key(row):
        date_val = row.get("Date", "")
        if pd.notna(date_val) and date_val != "":
            try:
                date_str = pd.Timestamp(date_val).strftime("%Y-%m-%d")
            except Exception:
                date_str = str(date_val)
        else:
            date_str = ""
        return (
            str(row.get("Winner","")).lower(),
            str(row.get("Loser","")).lower(),
            str(row.get("Round","")).lower(),
            str(row.get("Tournament","")).lower(),
            date_str,
        )

    existing_keys = {match_key(row) for _, row in df_existing_results.iterrows()}
    filtered = [m for m in new_matches if match_key(m) not in existing_keys]
    info(f"Nowych wyników: {len(filtered)}")
    
    # Usuń stare wiersze Upcoming — zostaną odbudowane na nowo
    if "Comment" in df_existing.columns:
        df_existing_clean = df_existing[df_existing["Comment"] != "Upcoming"].copy()
    else:
        df_existing_clean = df_existing.copy()

    if filtered:
        df_new = pd.DataFrame(filtered)
        for col in EXCEL_COLS:
            if col not in df_new.columns: df_new[col] = np.nan
        df_new = df_new[EXCEL_COLS]
        for col in EXCEL_COLS:
            if col not in df_existing_clean.columns: df_existing_clean[col] = np.nan
        df_combined = pd.concat([df_existing_clean[EXCEL_COLS], df_new], ignore_index=True)
        ok(f"Dopisano {len(filtered)} wyników")
    else:
        for col in EXCEL_COLS:
            if col not in df_existing_clean.columns: df_existing_clean[col] = np.nan
        df_combined = df_existing_clean[EXCEL_COLS]
        warn("Brak nowych wyników")
    
    # Sortowanie chronologiczne
    df_combined["Date"] = pd.to_datetime(df_combined["Date"], errors="coerce")
    df_combined = df_combined.sort_values("Date").reset_index(drop=True)
    
    # Ranking
    ranking_df = compute_ranking(df_combined)
    
    # Dopisz WRank, LRank, WPts, LPts
    rank_map = {}
    for rank, row in ranking_df.iterrows():
        rank_map[row["Player"]] = (rank, row["Points"])
    
    df_combined["WRank"] = df_combined["Winner"].map(lambda n: rank_map.get(n, (np.nan,))[0])
    df_combined["LRank"] = df_combined["Loser"].map(lambda n: rank_map.get(n, (np.nan,))[0])
    df_combined["WPts"]  = df_combined["Winner"].map(lambda n: rank_map.get(n, (np.nan,np.nan))[1])
    df_combined["LPts"]  = df_combined["Loser"].map(lambda n: rank_map.get(n, (np.nan,np.nan))[1])
    
    # Format daty
    df_combined["Date"] = df_combined["Date"].dt.strftime("%Y-%m-%d")
    
    # ─── UPCOMING → koniec SHEET_DATA ───
    # Konwertuj upcoming na format EXCEL_COLS i dopisz na koniec df_combined
    if upcoming:
        up_rows = []
        for u in upcoming:
            d = u.get("Date")
            date_str = d.strftime("%Y-%m-%d") if pd.notna(d) else np.nan
            row = {col: np.nan for col in EXCEL_COLS}
            row.update({
                "Location":   u.get("Location", ""),
                "Tournament": u.get("Tournament", ""),
                "Date":       date_str,
                "Series":     u.get("Series", ""),
                "Court":      u.get("Court", ""),
                "Surface":    u.get("Surface", ""),
                "Round":      u.get("Round", ""),
                "Best of":    3,
                "Winner":     u.get("Player1", ""),
                "Loser":      u.get("Player2", ""),
                "AvgW":       u.get("Odds1", np.nan),
                "AvgL":       u.get("Odds2", np.nan),
                "Comment":    "Upcoming",
            })
            up_rows.append(row)
        df_up = pd.DataFrame(up_rows)[EXCEL_COLS]
        df_combined = pd.concat([df_combined, df_up], ignore_index=True)
        ok(f"Dopisano {len(df_up)} nierozegranych meczów na koniec arkusza {SHEET_DATA}")
    
    # ─── UPCOMING → EXCEL ───
    if upcoming:
        up_df = pd.DataFrame(upcoming)
        up_df["Date"] = pd.to_datetime(up_df["Date"], errors="coerce").dt.strftime("%Y-%m-%d")
        for col in UPCOMING_COLS:
            if col not in up_df.columns: up_df[col] = np.nan
        up_df = up_df[UPCOMING_COLS]
        ok(f"Upcoming: {len(up_df)} meczów do zapisu")
    else:
        up_df = pd.DataFrame(columns=UPCOMING_COLS)
        warn("Brak upcoming")
    
    # Zapis
    try:
        with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a",
                            if_sheet_exists="replace") as writer:
            df_combined.to_excel(writer, sheet_name=SHEET_DATA, index=False)
            ranking_df.to_excel(writer, sheet_name=SHEET_RANKING, index=True)
            up_df.to_excel(writer, sheet_name=SHEET_UPCOMING, index=False)
        ok(f"Zapisano: {SHEET_DATA}={len(df_combined)} | {SHEET_RANKING}={len(ranking_df)} | {SHEET_UPCOMING}={len(up_df)}")
    except PermissionError:
        err("ZAMKNIJ EXCEL!"); return None, None
    
    try:
        wb = load_workbook(excel_path)
        hf = PatternFill("solid", fgColor="1A2B4A")
        hn = Font(bold=True, color="FFFFFF")
        for sn in [SHEET_DATA, SHEET_RANKING, SHEET_UPCOMING]:
            if sn in wb.sheetnames:
                ws = wb[sn]
                for c in ws[1]: c.fill, c.font = hf, hn
                ws.freeze_panes = "A2"
        wb.save(excel_path)
    except: pass
    
    return df_combined, ranking_df


def update_excel_from_tournament_url(excel_path, url, debug=False, dry_run=False):
    excel_path = Path(excel_path)
    if not excel_path.exists():
        err(f"Brak: {excel_path}")
        return None, None

    df_existing = pd.read_excel(excel_path, sheet_name=SHEET_DATA)
    name_idx = build_name_index(df_existing)
    soup = fetch_page(url)
    meta = parse_tournament_meta(soup, url)

    if debug:
        for i, table in enumerate(soup.find_all("table")):
            header = table.find("tr")
            if header:
                cols = [_clean(c.get_text()) for c in header.find_all(["th", "td"])]
                ttype = classify_table(table)
                print(f"Tabela {i}: [{ttype or '?'}] {cols}")
        return None, None

    results, upcoming = scrape_all(soup, meta)
    if dry_run:
        return pd.DataFrame(results), pd.DataFrame(upcoming)
    return write_excel(excel_path, results, upcoming, name_idx)


def update_excel_from_current_tournaments(excel_path, include_challengers=False, limit=None):
    excel_path = Path(excel_path)
    urls = discover_current_tournament_urls(limit=limit, include_challengers=include_challengers)
    info(f"Auto-discovered ATP URLs: {len(urls)}")

    if not urls:
        warn("Brak turniejów do przetworzenia")
        return {"urls": urls, "combined": None, "ranking": None}

    if not Path(excel_path).exists():
        err(f"Brak: {excel_path}")
        return {"urls": urls, "combined": None, "ranking": None}

    # Zbierz wyniki i upcoming ze WSZYSTKICH turniejów naraz
    df_existing = pd.read_excel(excel_path, sheet_name=SHEET_DATA)
    name_idx = build_name_index(df_existing)
    all_results = []
    all_upcoming = []

    for url in urls:
        hdr(f"LIVE UPDATE: {url}")
        try:
            soup = fetch_page(url)
            meta = parse_tournament_meta(soup, url)
            results, upcoming = scrape_all(soup, meta)
            all_results.extend(results)
            all_upcoming.extend(upcoming)
            ok(f"  {meta['tournament']}: {len(results)} wyników, {len(upcoming)} upcoming")
        except Exception as e:
            err(f"  Błąd przy {url}: {e}")

    info(f"Łącznie ze wszystkich turniejów: {len(all_results)} wyników, {len(all_upcoming)} upcoming")

    # Jeden zapis z danymi ze wszystkich turniejów
    combined_df, ranking_df = write_excel(excel_path, all_results, all_upcoming, name_idx)
    return {"urls": urls, "combined": combined_df, "ranking": ranking_df}


# ─── MAIN ───────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser()
    p.add_argument("--url", default=DEFAULT_URL)
    p.add_argument("--excel", default=DEFAULT_EXCEL)
    p.add_argument("--dry-run", action="store_true")
    p.add_argument("--debug", action="store_true")
    p.add_argument("--auto", action="store_true")
    p.add_argument("--include-challengers", action="store_true")
    p.add_argument("--limit", type=int, default=None)
    args = p.parse_args()
    
    excel_path = Path(args.excel)
    
    print(f"\n{Fore.CYAN}{Style.BRIGHT}")
    print("  ╔═══════════════════════════════════════╗")
    print("  ║   🎾  Tennis Scraper v4.2  🎾         ║")
    print("  ╚═══════════════════════════════════════╝")
    print(Style.RESET_ALL)
    
    hdr("1. Plik Excel")
    if not excel_path.exists():
        err(f"Brak: {excel_path}"); sys.exit(1)
    ok(f"OK: {excel_path}")
    
    df_existing = pd.read_excel(excel_path, sheet_name=SHEET_DATA)
    name_idx = build_name_index(df_existing)
    ok(f"Graczy: {len(name_idx)}")
    
    if args.auto:
        hdr("2. Auto-wyszukiwanie turniejow")
        urls = discover_current_tournament_urls(limit=args.limit, include_challengers=args.include_challengers)
        for url in urls:
            info(url)
        if args.dry_run:
            ok(f"Znaleziono: {len(urls)} URL")
            return
        update_excel_from_current_tournaments(excel_path, include_challengers=args.include_challengers, limit=args.limit)
        ok("Zaktualizowano live wyniki z auto-discovery")
        return

    hdr("2. Pobieranie")
    info(f"URL: {args.url}")
    soup = fetch_page(args.url)
    ok("Pobrano")
    
    meta = parse_tournament_meta(soup, args.url)
    info(f"{meta['tournament']} | {meta['series']}")
    
    if args.debug:
        hdr("DEBUG - tabele na stronie")
        for i, table in enumerate(soup.find_all("table")):
            header = table.find("tr")
            if header:
                cols = [_clean(c.get_text()) for c in header.find_all(["th","td"])]
                ttype = classify_table(table)
                print(f"\n  Tabela {i}: [{ttype or '?'}] {cols}")
                for j, row in enumerate(table.find_all("tr")[1:6]):
                    cells = [_clean(c.get_text())[:20] for c in row.find_all(["td","th"])]
                    links = [_clean(l.get_text())[:30] for l in row.find_all("a")]
                    print(f"    [{j}] {cells}")
                    if links: print(f"         links: {links}")
        return
    
    hdr("3. Parsowanie")
    results, upcoming = scrape_all(soup, meta)
    ok(f"Zakończone: {len(results)}")
    ok(f"Nadchodzące: {len(upcoming)}")
    
    if results:
        hdr("4. Zakończone")
        for r in results[:12]:
            d = r["Date"].strftime("%Y-%m-%d") if pd.notna(r["Date"]) else "????"
            rnd = (r["Round"] or "?")[:12]
            w = (r["Winner"] or "?")[:18]
            l = (r["Loser"] or "?")[:18]
            s = f"{r['Wsets']}-{r['Lsets']}"
            c = f" [{r['Comment']}]" if r['Comment'] != "Completed" else ""
            print(f"  {d} | {rnd:12} | {w:18} def. {l:18} ({s}){c}")
        if len(results) > 12: info(f"... +{len(results)-12}")
    
    if upcoming:
        hdr("5. Nadchodzące")
        for u in upcoming:
            d = u["Date"].strftime("%Y-%m-%d") if pd.notna(u["Date"]) else "????"
            rnd = (u["Round"] or "?")[:14]
            p1 = (u["Player1"] or "?")[:20]
            p2 = (u["Player2"] or "?")[:20]
            o = ""
            if not np.isnan(u.get("Odds1", np.nan)):
                o = f"[{u['Odds1']:.2f} / {u['Odds2']:.2f}]"
            print(f"  {d} | {rnd:14} | {p1:20} vs {p2:20} {o}")
    
    if args.dry_run:
        hdr("DRY RUN"); return
    
    hdr("6. Zapis")
    df_combined, ranking_df = write_excel(excel_path, results, upcoming, name_idx)
    if df_combined is None: return
    
    hdr("7. Top 15 Ranking")
    print(f"\n  {'#':>4}  {'Player':<22}  {'Pts':>6}  {'W':>3}  {'L':>3}  {'Win%':>6}  {'Best'}")
    print(f"  {'─'*4}  {'─'*22}  {'─'*6}  {'─'*3}  {'─'*3}  {'─'*6}  {'─'*15}")
    for rank, row in ranking_df.head(15).iterrows():
        medal = {1:"🥇",2:"🥈",3:"🥉"}.get(rank, "  ")
        print(f"  {rank:>4}  {medal}{row['Player']:<20}  {row['Points']:>6}  "
              f"{row['W']:>3}  {row['L']:>3}  {row['Win%']:>5}%  {row['BestResult']}")
    
    print(f"\n{Fore.GREEN}{Style.BRIGHT}✔ Gotowe! {len(df_combined)} meczów, "
          f"{len(upcoming)} nadchodzących.{Style.RESET_ALL}\n")


if __name__ == "__main__":
    main()
